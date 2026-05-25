#!/usr/bin/env python3
"""
Convert old IFR Excel format (single 'data' sheet) to new multi-sheet format.

Usage:
    python convert_ifr_format.py <old_file.xlsx> <template_file.xlsx> <output_file.xlsx>

Arguments:
    old_file       Old-format Excel file (e.g. Matthias_AT_befuellt_v2_IFR.xlsx)
    template_file  New-format example file used as structural template
                   (e.g. AT_20260520_V3_Claude.xlsx or FR_20260522_V1.1_Claude.xlsx)
    output_file    Path for the converted output file

The script copies the template, then overwrites data cells with values from the
old-format file, mapping them by row code. The export sheet is never modified.

Column mapping (both old and new formats):
    Col A: row code | Col B: label | Col C=aa=All | Col D=nf=Native Female
    Col E=nm=Native Male | Col F=mf=Migrant Female | Col G=mm=Migrant Male
"""

import sys
import shutil
import argparse
import re
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

COLS = ['aa', 'nf', 'nm', 'mf', 'mm']
# openpyxl column indices (1-based): C=3=aa, D=4=nf, E=5=nm, F=6=mf, G=7=mm
COL_IDX = {'aa': 3, 'nf': 4, 'nm': 5, 'mf': 6, 'mm': 7}


# ---------------------------------------------------------------------------
# Load old format
# ---------------------------------------------------------------------------

def load_old(filepath):
    """Return (data_by_code, imm_yearly, emi_yearly).

    data_by_code: {code: {col: value}}  – keyed on col A row code
    imm_yearly:   {year_int: {col: value}}  – 'Total migration in YYYY' rows
    emi_yearly:   {year_int: {col: value}}  – 'Total emigration in YYYY' rows
    """
    df = pd.read_excel(filepath, sheet_name='data', header=None)

    data = {}
    for _, row in df.iterrows():
        raw_code = row[0]
        if pd.isna(raw_code):
            continue
        code = str(raw_code).strip()
        if not code:
            continue
        vals = {}
        for i, col in enumerate(COLS):
            v = row[i + 2]
            if pd.isna(v) or v == '':
                vals[col] = None
            else:
                try:
                    vals[col] = float(v)
                except (ValueError, TypeError):
                    vals[col] = v  # keep strings (e.g. country name, source text)
        data[code] = vals

    imm_yearly = {}
    emi_yearly = {}
    for _, row in df.iterrows():
        label = str(row[1]).strip() if pd.notna(row[1]) else ''
        if label.startswith('Total migration in '):
            try:
                year = int(label.replace('Total migration in ', '').strip())
            except ValueError:
                continue
            vals = {}
            for i, col in enumerate(COLS):
                v = row[i + 2]
                try:
                    vals[col] = float(v) if pd.notna(v) and v != '' else None
                except (ValueError, TypeError):
                    vals[col] = None
            imm_yearly[year] = vals
        elif label.startswith('Total emigration in '):
            try:
                year = int(label.replace('Total emigration in ', '').strip())
            except ValueError:
                continue
            vals = {}
            for i, col in enumerate(COLS):
                v = row[i + 2]
                try:
                    vals[col] = float(v) if pd.notna(v) and v != '' else None
                except (ValueError, TypeError):
                    vals[col] = None
            emi_yearly[year] = vals

    return data, imm_yearly, emi_yearly


# ---------------------------------------------------------------------------
# Helpers for writing to openpyxl worksheet
# ---------------------------------------------------------------------------

def build_code_map(ws):
    """Return {code_str: row_number} from column A of a worksheet."""
    code_map = {}
    for row in ws.iter_rows():
        cell_a = row[0]
        if cell_a.value is not None:
            code_map[str(cell_a.value).strip()] = cell_a.row
    return code_map


def build_migration_year_maps(template_file):
    """Return (imm_year_map, emi_year_map) mapping year→openpyxl_row for immigration/emigration.

    Uses pandas (cached values) instead of openpyxl to handle ArrayFormula year labels.
    openpyxl row = pandas 0-index + 1.
    """
    df = pd.read_excel(template_file, sheet_name='migration', header=None)
    imm_year_map = {}
    emi_year_map = {}
    current_section = None

    for pandas_row_idx, row in df.iterrows():
        label = str(row[1]).strip() if pd.notna(row[1]) else ''
        if 'Immigration' in label:
            current_section = 'immigration'
        elif 'Emigration' in label:
            current_section = 'emigration'
        else:
            try:
                year = int(float(label))
                if 1900 <= year <= 2100 and current_section:
                    openpyxl_row = pandas_row_idx + 1
                    if current_section == 'immigration':
                        imm_year_map[year] = openpyxl_row
                    else:
                        emi_year_map[year] = openpyxl_row
            except (ValueError, TypeError):
                pass

    return imm_year_map, emi_year_map


def write_row(ws, row_num, vals_dict, cols=None):
    """Write values from vals_dict to the given row in ws."""
    if cols is None:
        cols = COL_IDX
    for col, col_idx in cols.items():
        v = vals_dict.get(col)
        if v is not None:
            ws.cell(row=row_num, column=col_idx).value = v


def write_scalar(ws, row_num, value, col='aa'):
    """Write a single scalar value to one column in ws."""
    if value is not None:
        ws.cell(row=row_num, column=COL_IDX[col]).value = value


def net_vals(old, code_m, code_e):
    """Return net {col: value} = old[code_m] - old[code_e]."""
    m_vals = old.get(code_m, {})
    e_vals = old.get(code_e, {})
    result = {}
    for col in COLS:
        m = m_vals.get(col)
        e = e_vals.get(col) or 0.0
        if m is not None:
            result[col] = m - e
    return result


# ---------------------------------------------------------------------------
# Per-sheet conversion logic
# ---------------------------------------------------------------------------

def convert_overview(ws, old):
    cm = build_code_map(ws)

    # Country identifiers
    for code in ('na', 'nn', 'i', 'is', 'ls'):
        if code in cm and code in old:
            write_scalar(ws, cm[code], old[code].get('aa'))

    # Generation date
    if 'nd' in cm:
        ws.cell(row=cm['nd'], column=COL_IDX['aa']).value = \
            datetime.now().strftime('%Y-%m-%d %H:%M')

    # Life expectancy – write all 5 columns
    if 'l' in cm and 'l' in old:
        write_row(ws, cm['l'], old['l'])


def convert_population(ws, old):
    cm = build_code_map(ws)

    if 'py' in cm and 'py' in old:
        write_scalar(ws, cm['py'], old['py'].get('aa'))

    for age in range(100):
        code = f'p{age:02d}'
        if code in cm and code in old:
            write_row(ws, cm[code], old[code])

    if 'pt' in cm and 'pt' in old:
        write_row(ws, cm['pt'], old['pt'])

    if 'ps' in cm and 'ps' in old:
        write_scalar(ws, cm['ps'], old['ps'].get('aa'))


def convert_births(ws, old):
    cm = build_code_map(ws)

    # Official fertility rate
    if 'f' in cm and 'f' in old:
        write_row(ws, cm['f'], old['f'])

    # Calculated births by fertility rate
    if 'f00' in cm and 'f00' in old:
        write_row(ws, cm['f00'], old['f00'])

    # Births source
    if 'fs' in cm and 'fs' in old:
        write_scalar(ws, cm['fs'], old['fs'].get('aa'))

    # b15 in new format = "15 years and younger" → cumulative sum of old b00–b15
    if 'b15' in cm:
        cumulative = {col: 0.0 for col in COLS}
        for age in range(16):
            code = f'b{age:02d}'
            if code in old:
                for col in COLS:
                    v = old[code].get(col)
                    if v is not None:
                        cumulative[col] += v
        non_zero = {col: v for col, v in cumulative.items() if v != 0.0}
        if non_zero:
            write_row(ws, cm['b15'], cumulative)

    # b16–b49: individual ages map 1-to-1
    for age in range(16, 50):
        code = f'b{age:02d}'
        if code in cm and code in old:
            write_row(ws, cm[code], old[code])


def convert_deaths(ws, old):
    cm = build_code_map(ws)

    for age in range(100):
        code = f'd{age:02d}'
        if code in cm and code in old:
            write_row(ws, cm[code], old[code])

    if 'ds' in cm and 'ds' in old:
        write_scalar(ws, cm['ds'], old['ds'].get('aa'))


def convert_migration(ws, old, imm_yearly, emi_yearly, template_file):
    cm = build_code_map(ws)
    imm_year_map, emi_year_map = build_migration_year_maps(template_file)

    # ma: net 10-year average (immigration avg – emigration avg)
    if 'ma' in cm and 'ma' in old:
        nv = net_vals(old, 'ma', 'ea')
        if nv:
            write_row(ws, cm['ma'], nv)

    # m00–m99: net migration by age (immigration – emigration)
    for age in range(100):
        code_m = f'm{age:02d}'
        code_e = f'e{age:02d}'
        if code_m in cm:
            nv = net_vals(old, code_m, code_e)
            if nv:
                write_row(ws, cm[code_m], nv)

    # Migration source
    if 'ms' in cm and 'ms' in old:
        write_scalar(ws, cm['ms'], old['ms'].get('aa'))

    # Yearly immigration – write to immigration section rows
    for year, vals in imm_yearly.items():
        if year in imm_year_map:
            write_row(ws, imm_year_map[year], vals)

    # Yearly emigration – write to emigration section rows
    for year, vals in emi_yearly.items():
        if year in emi_year_map:
            write_row(ws, emi_year_map[year], vals)


def convert_origin(ws, old):
    cm = build_code_map(ws)

    for code, vals in old.items():
        if code.startswith('c') and len(code) >= 2 and code in cm:
            write_row(ws, cm[code], vals)

    if 'cs' in cm and 'cs' in old:
        write_scalar(ws, cm['cs'], old['cs'].get('aa'))


# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------

def convert(old_file, template_file, output_file):
    print(f'Loading old format: {old_file}')
    old, imm_yearly, emi_yearly = load_old(old_file)
    print(f'  Loaded {len(old)} coded rows, '
          f'{len(imm_yearly)} immigration years, {len(emi_yearly)} emigration years')

    print(f'Copying template: {template_file} -> {output_file}')
    shutil.copy2(template_file, output_file)

    wb = load_workbook(output_file)
    print(f'  Sheets in template: {wb.sheetnames}')

    sheet_handlers = {
        'overview':  lambda ws: convert_overview(ws, old),
        'population': lambda ws: convert_population(ws, old),
        'births':    lambda ws: convert_births(ws, old),
        'deaths':    lambda ws: convert_deaths(ws, old),
        'migration': lambda ws: convert_migration(ws, old, imm_yearly, emi_yearly, template_file),
        'origin':    lambda ws: convert_origin(ws, old),
        # 'patria': no data in old format → leave template values
        # 'export': read-only → never modify
    }

    for sheet_name, handler in sheet_handlers.items():
        if sheet_name in wb.sheetnames:
            print(f'  Converting sheet: {sheet_name}')
            handler(wb[sheet_name])
        else:
            print(f'  Sheet not found in template: {sheet_name} – skipped')

    wb.save(output_file)
    print(f'Saved: {output_file}')


def main():
    parser = argparse.ArgumentParser(
        description='Convert old IFR Excel format to new multi-sheet format')
    parser.add_argument('old_file', nargs='?', default=None,
                        help='Old-format Excel file (single data sheet). If omitted, converts all old files in data/')
    parser.add_argument('template_file', nargs='?', default=None,
                        help='New-format template file (AT or FR example)')
    parser.add_argument('output_file', nargs='?', default=None,
                        help='Output path for converted Excel file')
    args = parser.parse_args()

    if args.old_file is None:
        convert_batch_in_data_folder()
    else:
        old_path = Path(args.old_file)
        tmpl_path = Path(args.template_file)
        out_path = Path(args.output_file)

        if not old_path.exists():
            print(f'Error: old file not found: {old_path}', file=sys.stderr)
            sys.exit(1)
        if not tmpl_path.exists():
            print(f'Error: template file not found: {tmpl_path}', file=sys.stderr)
            sys.exit(1)
        if out_path.exists():
            print(f'Warning: output file already exists, overwriting: {out_path}')

        convert(str(old_path), str(tmpl_path), str(out_path))


def extract_country_code_from_filename(filename: str) -> str:
    """Extract country code from filename like Matthias_AT.xlsx or Alex_FR.xlsx."""
    stem = Path(filename).stem
    match = re.search(r'_([A-Z]{2})$', stem)
    if match:
        return match.group(1).upper()
    match = re.search(r'([A-Z]{2})', stem)
    if match:
        return match.group(1).upper()
    return 'XX'


def find_template_for_country(country_code: str, data_dir: Path) -> Path:
    """Find a new-format template file for the given country code.
    
    Only uses original templates (AT_*.xlsx, FR_*.xlsx), not converted files.
    Falls back to AT template for countries without specific templates.
    """
    # Original templates only
    original_templates = {
        'AT': 'AT_20260520_V3_Claude.xlsx',
        'FR': 'FR_20260522_V1.1_Claude.xlsx'
    }
    
    if country_code in original_templates:
        tmpl_path = data_dir / original_templates[country_code]
        if tmpl_path.exists():
            return tmpl_path
    
    # Fallback: use AT template for any other country
    at_tmpl = data_dir / original_templates['AT']
    if at_tmpl.exists():
        return at_tmpl
    
    raise FileNotFoundError(f'No suitable template found for country {country_code} in {data_dir}')


def convert_batch_in_data_folder():
    """Convert all Matthias_* and Alex_* files in the data/ folder to new format."""
    data_dir = Path(__file__).resolve().parent / 'data'
    if not data_dir.exists():
        print(f'Error: data directory not found: {data_dir}', file=sys.stderr)
        sys.exit(1)

    print(f'Scanning data directory: {data_dir}')
    old_files = []
    for pattern in ['Matthias_*.xlsx', 'Alex_*.xlsx']:
        old_files.extend(data_dir.glob(pattern))
    old_files = [f for f in old_files if not f.name.startswith('~$')]
    old_files = sorted(old_files)

    if not old_files:
        print('No old-format files (Matthias_*.xlsx or Alex_*.xlsx) found in data/ folder')
        return

    print(f'Found {len(old_files)} old-format files to convert')
    today_str = datetime.now().strftime('%Y%m%d')

    for old_file in old_files:
        country_code = extract_country_code_from_filename(old_file.name)
        print(f'\nProcessing: {old_file.name} (country: {country_code})')

        try:
            template_file = find_template_for_country(country_code, data_dir)
            print(f'  Using template: {template_file.name}')
        except FileNotFoundError as e:
            print(f'  Error: {e} – skipping {old_file.name}')
            continue

        output_name = f'{country_code}_{today_str}_{old_file.name}'
        output_file = data_dir / output_name
        if output_file.exists():
            output_file.unlink()
            print(f'  Removed existing file: {output_name}')

        try:
            convert(str(old_file), str(template_file), str(output_file))
            print(f'  Successfully converted to: {output_name}')
        except Exception as e:
            print(f'  Error during conversion: {e}')
            if output_file.exists():
                output_file.unlink()



if __name__ == '__main__':
    main()
